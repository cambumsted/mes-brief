"""Build AMEX FY26 Q2 grid as an Excel workbook from the partnership-ideas.html source values."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(__file__).resolve().parent.parent / "downloads" / "AMEX-FY26-Q2-Grid.xlsx"

# Style helpers ---------------------------------------------------------------
ARIAL = "Arial"
HEADER_FILL = PatternFill("solid", start_color="091F2E")
HEADER_FONT = Font(name=ARIAL, bold=True, color="FFFFFF", size=10)
REGION_FILL = PatternFill("solid", start_color="E6F2FB")
REGION_FONT = Font(name=ARIAL, bold=True, color="091F2E", size=11)
MARKET_FONT = Font(name=ARIAL, color="000000", size=10)
INPUT_FONT  = Font(name=ARIAL, color="0000FF", size=10)   # blue = hardcoded input
FORMULA_FONT = Font(name=ARIAL, color="000000", size=10)   # black = formula
TOTAL_FILL = PatternFill("solid", start_color="091F2E")
TOTAL_FONT = Font(name=ARIAL, bold=True, color="FFFFFF", size=10)
THIN = Side(border_style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FMT_INT = '#,##0;(#,##0);-'
FMT_PCT = '0.00%;(0.00%);-'
FMT_USD = '$#,##0.00;($#,##0.00);-'
FMT_USD0 = '$#,##0;($#,##0);-'
FMT_MULT = '0.00"x"'

# Columns:
# A Region | B Market/Segment | C Audience | D Offer
# E Enrollments | F Redemptions | G Red Rate | H Red Spend (USD)
# I Offer Cost (USD) | J AOV Redeemers (USD) | K ROI | L Revenue (USD)
# M Source File | N Source Sheet | O Notes
HEADERS = [
    "Region", "Market / Segment", "Audience", "Offer",
    "Enrollments", "Redemptions", "Red. Rate", "Red. Spend (USD)",
    "Offer Cost (USD)", "AOV Redeemers (USD)", "ROI", "Revenue (USD)",
    "Source File", "Source Sheet", "Notes",
]

# Market rows — only hardcoded inputs. Region totals + derived metrics are formulas.
# Each market dict: keys for col E,F,H,I,J  + meta
EMEA_FILE = "EMEA_AMEX calc template.xlsx"
EMEA_SHEET = "AMEX FY26 Q2"
ASIA_FILE = "Asia_AMEX-Campaign-FY25-Q4.xlsx"
ASIA_SHEET = "FY26 Q2"
AMER_FILE = "Americas_FY26 - AMEX Performance Tracker.xlsx"
AMER_SHEET = "Key Metrics Worksheet"

DATA = [
    # ---- EMEA ----
    {"region": "EMEA", "is_total": True, "period": "Nov 11, 2025 – Jan 15, 2026",
     "market": "EMEA total", "audience": "", "offer": "Spend 500/100 (local) · 4 markets",
     "source_file": EMEA_FILE, "source_sheet": EMEA_SHEET, "notes": "AOV Redeemers used; EMEA also tracks AOV Enrolled (col G) — confirm canonical choice."},
    {"region": "EMEA", "market": "FR (cap 8,000)", "audience": "", "offer": "Spend €500, get €100",
     "enrollments": 4053, "redemptions": 33, "red_spend": 46837.46, "offer_cost": 3837.88, "aov": 1306.41,
     "source_file": EMEA_FILE, "source_sheet": EMEA_SHEET, "notes": "D16/H16/J16/M16/P16"},
    {"region": "EMEA", "market": "DE (cap 22,000)", "audience": "", "offer": "Spend €500, get €100",
     "enrollments": 17545, "redemptions": 286, "red_spend": 362022.14, "offer_cost": 33215.19, "aov": 1166.55,
     "source_file": EMEA_FILE, "source_sheet": EMEA_SHEET, "notes": "D25/H25/J25/M25/P25"},
    {"region": "EMEA", "market": "UK (cap 44,000)", "audience": "", "offer": "Spend £500, get £100",
     "enrollments": 45943, "redemptions": 733, "red_spend": 1636366.21, "offer_cost": 97060.91, "aov": 2077.12,
     "source_file": EMEA_FILE, "source_sheet": EMEA_SHEET, "notes": "D34/H34/J34/M34/P34 — UK exceeded cap (104%)"},
    {"region": "EMEA", "market": "IT (cap 8,000)", "audience": "", "offer": "Spend €500, get €100",
     "enrollments": 5414, "redemptions": 36, "red_spend": 80661.60, "offer_cost": 4190.42, "aov": 2062.00,
     "source_file": EMEA_FILE, "source_sheet": EMEA_SHEET, "notes": "D43/H43/J43/M43/P43"},

    # ---- ASIA ----
    {"region": "Asia", "is_total": True, "period": "Nov 18 – Dec 28, 2025",
     "market": "Asia total", "audience": "", "offer": "Mixed offers · 3 markets · 3 templates",
     "source_file": ASIA_FILE, "source_sheet": ASIA_SHEET,
     "notes": "Asia interim block has no enrollments column — enrollments derived = redemptions / red rate per market (each market hit cap)."},
    {"region": "Asia", "market": "AU Consumer (cap 20,000)", "audience": "", "offer": "Spend AUD$800, get AUD$200",
     "enrollments": 20000, "redemptions": 427, "red_spend": 504682, "offer_cost": 56141, "aov": 1181.93,
     "source_file": ASIA_FILE, "source_sheet": ASIA_SHEET, "notes": "Enrollments derived (cap reached). H13/D13/E13/F13"},
    {"region": "Asia", "market": "AU SBS (cap 10,000)", "audience": "", "offer": "Spend AUD$800, get AUD$200",
     "enrollments": 10000, "redemptions": 153, "red_spend": 266978, "offer_cost": 20226, "aov": 1744.95,
     "source_file": ASIA_FILE, "source_sheet": ASIA_SHEET, "notes": "Enrollments derived (cap reached). H14/D14/E14/F14"},
    {"region": "Asia", "market": "JP Consumer (cap 20,000)", "audience": "", "offer": "Spend ¥150,000, get ¥30,000",
     "enrollments": 20000, "redemptions": 310, "red_spend": 326214, "offer_cost": 59819, "aov": 1052.30,
     "source_file": ASIA_FILE, "source_sheet": ASIA_SHEET, "notes": "Enrollments derived (cap reached). H15/D15/E15/F15"},
    {"region": "Asia", "market": "JP SBS (cap 10,000)", "audience": "", "offer": "Spend ¥150,000, get ¥30,000",
     "enrollments": 10000, "redemptions": 58, "red_spend": 70530, "offer_cost": 11185, "aov": 1216.03,
     "source_file": ASIA_FILE, "source_sheet": ASIA_SHEET, "notes": "Enrollments derived (cap reached). H16/D16/E16/F16"},
    {"region": "Asia", "market": "SG All (cap 12,000)", "audience": "", "offer": "Spend S$800, get S$200",
     "enrollments": 12000, "redemptions": 48, "red_spend": 71691, "offer_cost": 7432, "aov": 1493.56,
     "source_file": ASIA_FILE, "source_sheet": ASIA_SHEET, "notes": "Enrollments derived (cap reached). H17/D17/E17/F17"},

    # ---- AMERICAS ----
    {"region": "Americas", "is_total": True, "period": "Nov 22, 2025 – Jan 17, 2026",
     "market": "Americas total", "audience": "", "offer": "Mixed offers · 4 segments concurrent",
     "source_file": AMER_FILE, "source_sheet": AMER_SHEET,
     "notes": "Americas calls this 'AOV' (redeemer-side) — equivalent to EMEA AOV Redeemers."},
    {"region": "Americas", "market": "Americas", "audience": "Consumer – No Spend", "offer": "Spend $70, get $15",
     "enrollments": 89990, "redemptions": 920, "red_spend": 154396.79, "offer_cost": 13800, "aov": 167.82,
     "source_file": AMER_FILE, "source_sheet": AMER_SHEET, "notes": "J10/J11/J12/J13/J15"},
    {"region": "Americas", "market": "Americas", "audience": "Consumer – SB Lookalike", "offer": "Spend $1,000, get $250",
     "enrollments": 22000, "redemptions": 124, "red_spend": 172204.43, "offer_cost": 31000, "aov": 1388.75,
     "source_file": AMER_FILE, "source_sheet": AMER_SHEET, "notes": "K10/K11/K12/K13/K15"},
    {"region": "Americas", "market": "Americas", "audience": "Consumer – Spend", "offer": "Spend $450, get $80",
     "enrollments": 20000, "redemptions": 118, "red_spend": 140393.81, "offer_cost": 9440, "aov": 1189.78,
     "source_file": AMER_FILE, "source_sheet": AMER_SHEET, "notes": "L10/L11/L12/L13/L15"},
    {"region": "Americas", "market": "Americas", "audience": "Small Business", "offer": "Spend $1,000, get $175",
     "enrollments": 24844, "redemptions": 61, "red_spend": 109727.61, "offer_cost": 10675, "aov": 1798.81,
     "source_file": AMER_FILE, "source_sheet": AMER_SHEET, "notes": "M10/M11/M12/M13/M15"},
]

# ----------------------------------------------------------------------------
wb = Workbook()
ws = wb.active
ws.title = "AMEX FY26 Q2"

# Title block
ws["A1"] = "AMEX FY26 Q2 grid — Nov 2025 – Jan 2026"
ws["A1"].font = Font(name=ARIAL, bold=True, size=14, color="091F2E")
ws.merge_cells("A1:O1")
ws["A2"] = ("Hardcoded inputs in blue; derived values in black formulas. "
            "Region totals SUM their child markets. AOV = Red Spend ÷ Redemptions; ROI = Red Spend ÷ Offer Cost.")
ws["A2"].font = Font(name=ARIAL, italic=True, color="7F7F7F", size=9)
ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells("A2:O2")
ws.row_dimensions[2].height = 28

HEADER_ROW = 4
# Header
for i, h in enumerate(HEADERS, start=1):
    c = ws.cell(row=HEADER_ROW, column=i, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
ws.row_dimensions[HEADER_ROW].height = 32

# Walk DATA, recording row indices for region totals so we can write SUM formulas after market rows are placed.
row = HEADER_ROW + 1
region_total_rows: dict[str, int] = {}      # region -> row of its total
region_market_rows: dict[str, list[int]] = {"EMEA": [], "Asia": [], "Americas": []}

for entry in DATA:
    region = entry["region"]
    is_total = entry.get("is_total", False)

    ws.cell(row=row, column=1, value=region)
    ws.cell(row=row, column=2, value=entry["market"])
    ws.cell(row=row, column=3, value=entry.get("audience", ""))
    ws.cell(row=row, column=4, value=entry["offer"])
    ws.cell(row=row, column=13, value=entry.get("source_file", ""))
    ws.cell(row=row, column=14, value=entry.get("source_sheet", ""))
    ws.cell(row=row, column=15, value=entry.get("notes", ""))
    ws.cell(row=row, column=12, value="TBD")   # Revenue placeholder

    if is_total:
        region_total_rows[region] = row
        # Placeholder note in market column with period
        ws.cell(row=row, column=2, value=f"{region} total — {entry['period']}")
        # Style row
        for col in range(1, 16):
            cell = ws.cell(row=row, column=col)
            cell.fill = REGION_FILL
            cell.font = REGION_FONT if col <= 4 else Font(name=ARIAL, bold=True, color="000000", size=10)
            cell.border = BORDER
    else:
        region_market_rows[region].append(row)
        ws.cell(row=row, column=5, value=entry["enrollments"])
        ws.cell(row=row, column=6, value=entry["redemptions"])
        ws.cell(row=row, column=8, value=entry["red_spend"])
        ws.cell(row=row, column=9, value=entry["offer_cost"])
        ws.cell(row=row, column=10, value=entry["aov"])  # AOV pulled from source, not derived

        # Derived formulas
        ws.cell(row=row, column=7, value=f"=IFERROR(F{row}/E{row},0)")     # Red Rate
        ws.cell(row=row, column=11, value=f"=IFERROR(H{row}/I{row},0)")    # ROI

        # Font coloring: inputs blue, formulas black
        for col in (5, 6, 8, 9, 10):
            ws.cell(row=row, column=col).font = INPUT_FONT
        for col in (7, 11):
            ws.cell(row=row, column=col).font = FORMULA_FONT
        for col in (1, 2, 3, 4, 12, 13, 14, 15):
            ws.cell(row=row, column=col).font = MARKET_FONT
        for col in range(1, 16):
            ws.cell(row=row, column=col).border = BORDER

    row += 1

# Fill in region total formulas
for region, total_row in region_total_rows.items():
    child_rows = region_market_rows[region]
    rngE = ",".join(f"E{r}" for r in child_rows)
    rngF = ",".join(f"F{r}" for r in child_rows)
    rngH = ",".join(f"H{r}" for r in child_rows)
    rngI = ",".join(f"I{r}" for r in child_rows)
    ws.cell(row=total_row, column=5, value=f"=SUM({rngE})")
    ws.cell(row=total_row, column=6, value=f"=SUM({rngF})")
    ws.cell(row=total_row, column=7, value=f"=IFERROR(F{total_row}/E{total_row},0)")
    ws.cell(row=total_row, column=8, value=f"=SUM({rngH})")
    ws.cell(row=total_row, column=9, value=f"=SUM({rngI})")
    ws.cell(row=total_row, column=10, value=f"=IFERROR(H{total_row}/F{total_row},0)")
    ws.cell(row=total_row, column=11, value=f"=IFERROR(H{total_row}/I{total_row},0)")

# Grand total row
grand_row = row + 1
ws.cell(row=grand_row, column=1, value="GRAND TOTAL")
ws.cell(row=grand_row, column=2, value="All regions")
all_total_rows = list(region_total_rows.values())
def total_sum(col_letter: str) -> str:
    return "=" + "+".join(f"{col_letter}{r}" for r in all_total_rows)
ws.cell(row=grand_row, column=5, value=total_sum("E"))
ws.cell(row=grand_row, column=6, value=total_sum("F"))
ws.cell(row=grand_row, column=7, value=f"=IFERROR(F{grand_row}/E{grand_row},0)")
ws.cell(row=grand_row, column=8, value=total_sum("H"))
ws.cell(row=grand_row, column=9, value=total_sum("I"))
ws.cell(row=grand_row, column=10, value=f"=IFERROR(H{grand_row}/F{grand_row},0)")
ws.cell(row=grand_row, column=11, value=f"=IFERROR(H{grand_row}/I{grand_row},0)")
ws.cell(row=grand_row, column=12, value="TBD")
for col in range(1, 16):
    cell = ws.cell(row=grand_row, column=col)
    cell.fill = TOTAL_FILL
    cell.font = TOTAL_FONT
    cell.border = BORDER

# Number formats for all data rows
data_rows = list(range(HEADER_ROW + 1, grand_row + 1))
for r in data_rows:
    ws.cell(row=r, column=5).number_format = FMT_INT
    ws.cell(row=r, column=6).number_format = FMT_INT
    ws.cell(row=r, column=7).number_format = FMT_PCT
    ws.cell(row=r, column=8).number_format = FMT_USD
    ws.cell(row=r, column=9).number_format = FMT_USD
    ws.cell(row=r, column=10).number_format = FMT_USD
    ws.cell(row=r, column=11).number_format = FMT_MULT

# Column widths
widths = {
    "A": 11, "B": 30, "C": 22, "D": 30,
    "E": 13, "F": 12, "G": 11, "H": 17,
    "I": 16, "J": 18, "K": 9, "L": 14,
    "M": 38, "N": 22, "O": 50,
}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# Wrap notes column
for r in data_rows:
    ws.cell(row=r, column=15).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=r, column=4).alignment = Alignment(wrap_text=True, vertical="top")

ws.freeze_panes = "C5"

# ============================================================================
# Sheet 2: Metric Definitions
ws2 = wb.create_sheet("Metric Definitions")
defs = [
    ("Region / Market", "Region or market/segment rollup for the AMEX campaign row."),
    ("Audience", "Audience segment within the region (used by Americas to break out Consumer No-Spend, SB Lookalike, Consumer Spend, and Small Business)."),
    ("Offer", "The offer construct presented to cardholders (spend threshold and reward)."),
    ("Enrollments", "Intent step: customer accepted/opted into the offer."),
    ("Redemptions", "Conversion step: qualifying purchase made by an enrolled customer."),
    ("Red. Rate", "Redemption Rate = Redemptions ÷ Enrollments."),
    ("Red. Spend (USD)", "Total customer spend in qualifying purchases that triggered the offer (USD)."),
    ("Offer Cost (USD)", "Cost to Microsoft of fulfilling offer rewards (USD)."),
    ("AOV Redeemers (USD)", "Average order value of redeeming customers (USD). EMEA reports this as 'AOV Redeemers'; Americas reports it as 'AOV'. Asia interim block has no AOV column — derived = Red Spend ÷ Redemptions."),
    ("ROI", "ROI = Redemption Spend ÷ Offer Cost."),
    ("Revenue (USD)", "TBD. Needs final definition (e.g., Redemption Spend only vs. broader attributed revenue)."),
]
ws2["A1"] = "AMEX FY26 Q2 — Metric Definitions"
ws2["A1"].font = Font(name=ARIAL, bold=True, size=14, color="091F2E")
ws2.merge_cells("A1:B1")

ws2["A3"] = "Metric"
ws2["B3"] = "Definition"
for c in (ws2["A3"], ws2["B3"]):
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BORDER

for i, (m, d) in enumerate(defs, start=4):
    ws2.cell(row=i, column=1, value=m).font = Font(name=ARIAL, bold=True, size=10)
    ws2.cell(row=i, column=2, value=d).font = Font(name=ARIAL, size=10)
    ws2.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws2.cell(row=i, column=1).border = BORDER
    ws2.cell(row=i, column=2).border = BORDER

ws2.column_dimensions["A"].width = 24
ws2.column_dimensions["B"].width = 95
for i in range(4, 4 + len(defs)):
    ws2.row_dimensions[i].height = 36

wb.save(OUT)
print(f"Wrote {OUT}")
